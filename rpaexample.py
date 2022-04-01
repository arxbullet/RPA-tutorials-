import rpa as r
import openpyxl
import time

 
# Se inicia la libreria rpa con el metodo init()
# Con esto se abre la ventana del navegador para poder realizar acciones
r.init()

# Se crea el objeto del libro excel para poder trabajar en el
# El método load_workbook de la libreria openpyxl recibe como parametro la ruta al archivo xlsx
wb = openpyxl.load_workbook('rpaexample.xlsx')

# Se obtiene la hoja activa del libro cargado anteriormente con el atributo active del objeto workbook(wb)
ws = wb.active

rows = ws.max_row

for i in range(2, rows + 1):
    text = ws.cell(row= i , column=2).value

    r.url('https://google.cl')
    
    r.type('/html/body/div[1]/div[3]/form/div[1]/div[1]/div[1]/div/div[2]/input', text, '[enter]')#copiar xpath
    
    time.sleep(5)

r.close()
